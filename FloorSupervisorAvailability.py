from openpyxl import load_workbook
from openpyxl import Workbook
import datetime as dt
import functions

# Writing to the file
#write = Workbook()
#ws = write.active


def ChangeTime(start, hours, minutes):
	pass

def main():

	""" Read Stage """
	now = dt.datetime.now()
	delta = dt.timedelta(hours=2)
	pgdelta = dt.timedelta(hours=2, minutes=30)

	wb = load_workbook('inputSchedules/Sept 15 Cal.xlsx', read_only=True)
	sheet = wb['Sept 2015']

	select = []

	for row in sheet.rows:
		if row[3].value:
			if 'P' in row[3].value and row[7].value == 'FIRM':
				select.append((row[0].value, row[1].value, row[2].value, row[3].value, row[7].value, row[9].value, row[10].value))

	""" Write Stage """
	write = Workbook()
	ws = write.active
	# Write the header
	ws['A1'] = "Day"
	ws['B1'] = "Date"
	ws['C1'] = "Venue"
	ws['D1'] = "E"
	ws['E1'] = "Status"
	ws['F1'] = "Time"
	ws['G1'] = "Event"
	ws['H1'] = "FS Initials"
	ws['I1'] = "Comments"

	for i in range(len(select)):
		ws['A'+str(i+2)] = select[i][0]
		ws['B'+str(i+2)] = select[i][1].date()
		ws['C'+str(i+2)] = select[i][2]
		ws['D'+str(i+2)] = select[i][3]
		ws['E'+str(i+2)] = select[i][4]
		if select[i][2] == 'P&G':
			ws['F'+str(i+2)] = (dt.datetime.combine(dt.date(1,1,1),select[i][5]) - pgdelta).time()
		else:
			ws['F'+str(i+2)] = (dt.datetime.combine(dt.date(1,1,1),select[i][5]) - delta).time()
		ws['G'+str(i+2)] = select[i][6]
		ws['I'+str(i+2)] = ''
	# Adjust column width for larger columns
	ws = functions.columnWidth(ws, [('B', 20), ('G', 40), ('I', 40)])

	# Hide columns for later
	ws.column_dimensions.group('D', 'E', hidden=True)
	
	# Save 
	write.save("FloorSupervisorAvailablitiy.xlsx")



if __name__ == "__main__":
    main()