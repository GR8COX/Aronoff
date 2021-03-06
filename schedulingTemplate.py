from openpyxl import load_workbook
from openpyxl import Workbook
import datetime as dt
import os, sys

# Writing to the file
write = Workbook()
ws = write.active

# Function for loading all the Floor supervisor availability
def loadWorkbook(path, form):
	wb = load_workbook((path+ "\\" + form), read_only=True)
	sheet = wb['Sheet'] 
	select = []
	for row in sheet.rows: 
		if row[7].value: 
			select.append((row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value))
	return select

#def type():
#	if time is after 5pm = 'N'
#	if time is before 5pm = 'M'
#	if there are two shows in the 'P&G' in one day = 'D'



def main():
	shifts = {}

	# Loop through and grab each Floor supervisor availability form 
	path = 'floorSupervisorAvailability'
	for template in os.listdir(path):
		# this loops through everything in the directory specified by 'path'
		if ('FloorSupervisorAvailablitiy_' in template) and ('~' not in template):
			initials = template[template.index('_') + 1: (template.index('_') + 3)]
			data = loadWorkbook(path, template)
			shifts[initials] = data
	print(shifts)
	

	# Open the existing Scheduling Template 
	modfile = 'schedulingTemplate/Scheduling template default.xlsm'
	wb = load_workbook(modfile)
	ws = wb.get_active_sheet()
	# write out to new file with supervisors initials added to appropriate row
	#write = Workbook()
	#aws = write.active
	# Write the headers for our volunteers
	
	#init = 'P'
	#for key, value in shifts:
	#	ws[init + '3'] = key
	#	init = chr(ord(init) + 1)


	# include logic for the columns type, number of shifts, availability for each day 


	# Sum of shifts

	# Sum of per day

	# Sum of remaining 

	# Save what we have written to file
	write.save("schedulingTemplate/Scheduling template default.xlsx")

if __name__ == "__main__":
    main()
